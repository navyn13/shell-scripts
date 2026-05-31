"""
Load master data source as rows (list of lists of strings).
Supports .csv and .xlsx so you can pass either file path.
"""
import csv
from pathlib import Path


def load_source_rows(path):
    """
    Load a CSV or Excel file and return rows as list[list[str]].
    path: Path or str to .csv or .xlsx file.
    """
    path = Path(path)
    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        return _load_xlsx(path)
    return _load_csv(path)


def _load_csv(path):
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        return list(reader)


def _load_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "Reading Excel (.xlsx) requires openpyxl. Install with: pip install openpyxl"
        ) from None

    # Need cell objects (not values_only) to access number_format,
    # so we can mimic what Excel actually displays to the user.
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=False):
        rows.append([_cell_str(c.value, c.number_format) for c in row])
    wb.close()

    if rows:
        max_len = max(len(r) for r in rows)
        for r in rows:
            while len(r) < max_len:
                r.append("")
    return rows


def _decimals_from_format(number_format):
    """
    Parse the number of decimal places from an Excel number_format string.
    Returns None if the format does not specify a fixed decimal precision
    (e.g. 'General', text formats), otherwise an int >= 0.

    Examples:
        '0'        -> 0   (display as integer)
        '0.00'     -> 2
        '#,##0.0'  -> 1
        '0.000'    -> 3
        'General'  -> None
        ''         -> None
    """
    if not number_format or number_format.lower() == "general":
        return None

    # Strip any text/quoted sections that don't affect numeric precision
    fmt = number_format.split(";")[0]

    if "." not in fmt:
        # No decimal point in format => Excel displays as an integer
        if any(c in fmt for c in "0#"):
            return 0
        return None

    decimal_part = fmt.split(".", 1)[1]
    decimals = 0
    for c in decimal_part:
        if c in "0#":
            decimals += 1
        elif c in ",":
            continue
        else:
            break
    return decimals


def _cell_str(value, number_format="General"):
    """
    Convert an Excel cell value to a clean string, applying the cell's
    displayed number format so the output matches what Excel shows.

    - None  -> ""
    - str   -> stripped string
    - int   -> str(int)
    - float -> formatted per number_format:
        * format '0'    -> rounded integer (e.g. 2776.657 -> '2777')
        * format '0.00' -> rounded to 2 decimals
        * 'General' or no precision -> preserve full value (drop trailing .0)
    """
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        decimals = _decimals_from_format(number_format)
        if decimals is None:
            # 'General' / unspecified precision: preserve full value
            return str(int(value)) if value.is_integer() else str(value)
        rounded = round(value, decimals)
        if decimals == 0:
            return str(int(rounded))
        # Use fixed-precision formatting then strip trailing zeros if not meaningful
        return f"{rounded:.{decimals}f}"
    return str(value)
